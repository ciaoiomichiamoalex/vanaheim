from calendar import Calendar
from datetime import date
from openpyxl.styles import Alignment, Font
from share import sqlmng
from share.common import logger_ini
import openpyxl
import os

PATH_PRJ = 'c:/source/vanaheim'
PATH_LOG = f"{PATH_PRJ}/log/vanaheim_{date.today().strftime('%Y_%m_%d')}.log"
PATH_RES = f'{PATH_PRJ}/res'
PATH_SCHEME = f'{PATH_PRJ}/scheme'

QUERY_OVERVIEW_DATA = """
    SELECT numero_documento,
        data_documento,
        ragione_sociale,
        sede_consegna,
        quantita,
        data_consegna,
        targa
    FROM vanaheim.consegne
    WHERE EXTRACT(YEAR FROM data_consegna) = ?
        AND EXTRACT(MONTH FROM data_consegna) = ?
    ORDER BY numero_documento;
"""
QUERY_SUMMARY_VIAGGI = """
    WITH viaggi AS (
        SELECT ROW_NUMBER() OVER (PARTITION BY targa ORDER BY data_consegna, sede_consegna) AS id,
            targa,
            data_consegna, 
            sede_consegna
        FROM vanaheim.consegne
        WHERE EXTRACT(YEAR FROM data_documento) = ?
    ) 
    SELECT v1.data_consegna data_es745wh,
        v1.sede_consegna sede_es745wh,
        NULL gap,
        v2.sede_consegna sede_fc065zw,
        v2.data_consegna data_fc065zw
    FROM (
        SELECT id,
            data_consegna,
            sede_consegna
        FROM viaggi
        WHERE targa = 'ES745WH'
    ) v1
        FULL JOIN (
            SELECT id, 
                data_consegna,
                sede_consegna
            FROM viaggi
            WHERE targa = 'FC065ZW'
        ) v2 ON v1.id = v2.id
    ORDER BY COALESCE(v1.id, v2.id);
"""

DEFAULT_FONT = 'Arial'
FORMATS = {
    type(None): 'General',
    str: '@',
    int: '#,##0',
    date: 'dd/mm',
    'number': '0'
}


def overview_gnr(anno: int = date.today().year, mese: int = date.today().month) -> None:
    """Generate the month overview excel by taking data from the database.

    :param int anno: The desired year for the overview.
    :param int mese: The desired month for the overview.
    """
    logger = logger_ini(PATH_LOG, 'overview_doc')
    cursor = sqlmng.conx_ini()

    consegne = sqlmng.conx_read(cursor, QUERY_OVERVIEW_DATA, (anno, mese)).fetchall()
    if not consegne:
        logger.warning(f'no record founded in {anno}/{mese:0>2}... skipping overview!')
        return

    wb = openpyxl.load_workbook(f'{PATH_RES}/{anno}_{mese:0>2}.xlsx' if os.path.isfile(f'{PATH_RES}/{anno}_{mese:0>2}.xlsx') else f'{PATH_SCHEME}/consegne.xlsx')
    ws = wb['consegne']
    for row_num, row in enumerate(consegne, start=2):
        for col_num, col in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num).value = col
            ws.cell(row=row_num, column=col_num).font = Font(name=DEFAULT_FONT)
            ws.cell(row=row_num, column=col_num).number_format = FORMATS[type(col)] if col_num != 1 else FORMATS['number']

    sheets = [
        wb['cifre'],
        wb['litri'],
        wb['cifre manuale'],
        wb['litri manuale']
    ]

    for ws in sheets:
        ws.cell(row=1, column=1).value = date.today().year
        ws.cell(row=1, column=1).font = Font(name=DEFAULT_FONT)
        ws.cell(row=1, column=1).number_format = FORMATS['number']

        for row_num, day in enumerate([d for d in Calendar().itermonthdates(anno, mese) if d.month == mese], start=3):
            ws.cell(row=row_num, column=1).value = day
            ws.cell(row=row_num, column=1).font = Font(name=DEFAULT_FONT)
            ws.cell(row=row_num, column=1).number_format = FORMATS[date]
        # TODO: remove extra days (31/06, 30/02, ...) => openpyxl non aggiorna le formule

    logger.info('saving overview for %(data)s... [%(data)s.xlsx]' % {'data': f'{anno}_{mese:0>2}'})
    wb.save(f'{PATH_RES}/{anno}_{mese:0>2}.xlsx')
    cursor.close()


def summary_viaggi(anno: int = date.today().year) -> None:
    """Generate the year trips summary excel by taking data from the database.

    :param int anno: The desired year for the summary.
    """
    logger = logger_ini(PATH_LOG, 'overview_doc')
    cursor = sqlmng.conx_ini()

    viaggi = sqlmng.conx_read(cursor, QUERY_SUMMARY_VIAGGI, [anno]).fetchall()
    if not viaggi:
        logger.warning(f'no record founded in {anno}... skipping summary!')
        return

    wb = openpyxl.load_workbook(f'{PATH_SCHEME}/viaggi.xlsx')
    ws = wb['viaggi']
    for row_num, row in enumerate(viaggi, start=3):
        for col_num, col in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num).value = col
            ws.cell(row=row_num, column=col_num).font = Font(name=DEFAULT_FONT)
            ws.cell(row=row_num, column=col_num).number_format = FORMATS[type(col)]

            if col_num == 4:
                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='right')

    logger.info('saving summary for %(anno)d... [%(anno)d_TRIPS.xlsx]' % {'anno': anno})
    wb.save(f'{PATH_RES}/{anno}_TRIPS.xlsx')
    cursor.close()


if __name__ == '__main__':
    overview_gnr()
    summary_viaggi()
