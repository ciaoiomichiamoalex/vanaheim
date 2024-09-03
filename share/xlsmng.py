from openpyxl.styles import Font
import datetime
import openpyxl

DEFAULT_FONT = 'Aptos Narrow'
FORMATS = {
    type(None): 'General',
    str: '@',
    int: '#,##0',
    float: '#,##0.00',
    datetime.date: 'dd/mm/yyyy',
    datetime.time: 'h:mm:ss;@',
    datetime.datetime: 'dd/mm/yyyy h:mm:ss;@'
}


def write_excel(fou: str, rows: list, sheet_name: str = None, header: list[str] = None, font_face: str = DEFAULT_FONT) -> None:
    """Write a list of values into Excel file.

    :param str fou: Path to the result file.
    :param list rows: List of values to write.
    :param str sheet_name: Name of the sheet into the file, defaults to None.
    :param list[str] header: List of column names, defaults to None.
    :param str font_face: Font face of the file, defaults to DEFAULT_FONT constant.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if sheet_name:
        ws.title = sheet_name

    if header:
        for col_num, name in enumerate(header, start=1):
            ws.cell(row=1, column=col_num).value = name
            ws.cell(row=1, column=col_num).font = Font(name=font_face, bold=True)
            ws.cell(row=1, column=col_num).number_format = FORMATS[type(name)]

    for row_num, row in enumerate(rows, start=2 if header else 1):
        for col_num, col in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num).value = col
            ws.cell(row=row_num, column=col_num).font = Font(name=font_face)
            ws.cell(row=row_num, column=col_num).number_format = FORMATS[type(col)]

    if header:
        ws.auto_filter.ref = ws.dimensions
    wb.save(fou)
